import pandas as pd
import wbgapi as wb
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime

print("📦 Fetching live data from World Bank API...")

# ── 1. PULL DATA ────────────────────────────────────────────────────────────

# GDP (current US$)
gdp = wb.data.DataFrame("NY.GDP.MKTP.CD", time=range(2013, 2023), labels=True).reset_index()
gdp.columns = [str(c).replace("YR", "") for c in gdp.columns]
gdp = gdp.rename(columns={"economy": "Country Code", "Country": "Country"})

# GDP Growth (%)
growth = wb.data.DataFrame("NY.GDP.MKTP.KD.ZG", time=range(2013, 2023), labels=True).reset_index()
growth.columns = [str(c).replace("YR", "") for c in growth.columns]

# GDP Per Capita
per_capita = wb.data.DataFrame("NY.GDP.PCAP.CD", time=range(2013, 2023), labels=True).reset_index()
per_capita.columns = [str(c).replace("YR", "") for c in per_capita.columns]

print("✅ Data fetched successfully!")

# ── 2. CLEAN & ANALYZE ──────────────────────────────────────────────────────

year_cols = [str(y) for y in range(2013, 2023)]

# Latest year GDP (2022)
latest = gdp[["Country", "2022"]].dropna().sort_values("2022", ascending=False)
latest.columns = ["Country", "GDP 2022 (USD)"]
latest["GDP 2022 (Trillion)"] = (latest["GDP 2022 (USD)"] / 1e12).round(2)
top10 = latest.head(10).reset_index(drop=True)

# Global KPIs
total_world_gdp = latest["GDP 2022 (USD)"].sum()
avg_gdp = latest["GDP 2022 (USD)"].mean()
top_country = top10.iloc[0]["Country"]
top_gdp = top10.iloc[0]["GDP 2022 (Trillion)"]

# GDP Growth latest year
growth_latest = growth[["Country", "2022"]].dropna().sort_values("2022", ascending=False)
growth_latest.columns = ["Country", "GDP Growth % (2022)"]
growth_latest["GDP Growth % (2022)"] = growth_latest["GDP Growth % (2022)"].round(2)

# Per capita latest
pc_latest = per_capita[["Country", "2022"]].dropna().sort_values("2022", ascending=False)
pc_latest.columns = ["Country", "GDP Per Capita 2022 (USD)"]
pc_latest["GDP Per Capita 2022 (USD)"] = pc_latest["GDP Per Capita 2022 (USD)"].round(0)

print("✅ Analysis complete!")

# ── 3. BUILD EXCEL REPORT ───────────────────────────────────────────────────

today = datetime.today().strftime("%Y-%m-%d")
filename = f"GDP_Economic_Report_{today}.xlsx"

wb_excel = openpyxl.Workbook()

# Colors
DARK_BLUE = "1F4E79"
MID_BLUE  = "2E75B6"
LIGHT_BLUE = "DEEAF1"
WHITE = "FFFFFF"
GRAY = "F2F2F2"

def style_header(cell, bg=DARK_BLUE, font_color=WHITE, size=11, bold=True):
    cell.font = Font(bold=bold, color=font_color, size=size, name="Arial")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_cell(cell, bold=False, color=None, align="left", size=10):
    cell.font = Font(bold=bold, color=color or "000000", size=size, name="Arial")
    cell.alignment = Alignment(horizontal=align, vertical="center")

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# ── SHEET 1: Executive Summary ──────────────────────────────────────────────
ws1 = wb_excel.active
ws1.title = "Executive Summary"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 35
ws1.column_dimensions["B"].width = 30

# Title
ws1.merge_cells("A1:B1")
ws1["A1"] = "🌍 Global GDP Economic Report"
ws1["A1"].font = Font(bold=True, size=16, color=DARK_BLUE, name="Arial")
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 35

ws1.merge_cells("A2:B2")
ws1["A2"] = f"Generated: {datetime.today().strftime('%B %d, %Y')}  |  Source: World Bank API"
ws1["A2"].font = Font(size=10, color="888888", name="Arial")
ws1["A2"].alignment = Alignment(horizontal="center")
ws1.row_dimensions[2].height = 20

# KPI Headers
ws1.row_dimensions[4].height = 25
for col, header in enumerate(["Indicator", "Value"], 1):
    cell = ws1.cell(row=4, column=col, value=header)
    style_header(cell)

# KPI Data
kpis = [
    ("Total World GDP (2022)", f"${total_world_gdp/1e12:,.2f} Trillion"),
    ("Average Country GDP (2022)", f"${avg_gdp/1e9:,.2f} Billion"),
    ("Largest Economy (2022)", f"{top_country} — ${top_gdp}T"),
    ("Number of Countries Tracked", f"{len(latest):,}"),
    ("Data Period", "2013 – 2022"),
]
for i, (k, v) in enumerate(kpis, 5):
    ws1.row_dimensions[i].height = 22
    c1 = ws1.cell(row=i, column=1, value=k)
    c2 = ws1.cell(row=i, column=2, value=v)
    bg = GRAY if i % 2 == 0 else WHITE
    for c in [c1, c2]:
        c.fill = PatternFill("solid", fgColor=bg)
        c.border = thin_border()
        style_cell(c, align="center")

# ── SHEET 2: Top 10 Economies ───────────────────────────────────────────────
ws2 = wb_excel.create_sheet("Top 10 Economies")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 25

ws2.merge_cells("A1:C1")
ws2["A1"] = "Top 10 Economies by GDP (2022)"
ws2["A1"].font = Font(bold=True, size=14, color=DARK_BLUE, name="Arial")
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

for col, header in enumerate(["Rank", "Country", "GDP (Trillion USD)"], 1):
    cell = ws2.cell(row=2, column=col, value=header)
    style_header(cell)
ws2.row_dimensions[2].height = 25

for i, row in top10.iterrows():
    r = i + 3
    ws2.row_dimensions[r].height = 22
    bg = LIGHT_BLUE if i % 2 == 0 else WHITE
    for col, val in enumerate([i+1, row["Country"], row["GDP 2022 (Trillion)"]], 1):
        cell = ws2.cell(row=r, column=col, value=val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = thin_border()
        style_cell(cell, align="center")

# Add bar chart
chart2 = BarChart()
chart2.type = "bar"
chart2.title = "Top 10 Economies by GDP (2022)"
chart2.y_axis.title = "GDP (Trillion USD)"
chart2.x_axis.title = "Country"
chart2.style = 10
chart2.width = 18
chart2.height = 12
data_ref = Reference(ws2, min_col=3, min_row=2, max_row=12)
cats_ref = Reference(ws2, min_col=2, min_row=3, max_row=12)
chart2.add_data(data_ref, titles_from_data=True)
chart2.set_categories(cats_ref)
ws2.add_chart(chart2, "E2")

# ── SHEET 3: GDP Growth ─────────────────────────────────────────────────────
ws3 = wb_excel.create_sheet("GDP Growth 2022")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 25

ws3.merge_cells("A1:B1")
ws3["A1"] = "Top 20 Countries by GDP Growth Rate (2022)"
ws3["A1"].font = Font(bold=True, size=14, color=DARK_BLUE, name="Arial")
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

for col, header in enumerate(["Country", "GDP Growth % (2022)"], 1):
    cell = ws3.cell(row=2, column=col, value=header)
    style_header(cell)

top20_growth = growth_latest.head(20).reset_index(drop=True)
for i, row in top20_growth.iterrows():
    r = i + 3
    ws3.row_dimensions[r].height = 22
    bg = LIGHT_BLUE if i % 2 == 0 else WHITE
    for col, val in enumerate([row["Country"], row["GDP Growth % (2022)"]], 1):
        cell = ws3.cell(row=r, column=col, value=val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = thin_border()
        style_cell(cell, align="center")

# ── SHEET 4: GDP Per Capita ─────────────────────────────────────────────────
ws4 = wb_excel.create_sheet("GDP Per Capita")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 30
ws4.column_dimensions["B"].width = 25

ws4.merge_cells("A1:B1")
ws4["A1"] = "Top 20 Countries by GDP Per Capita (2022)"
ws4["A1"].font = Font(bold=True, size=14, color=DARK_BLUE, name="Arial")
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 30

for col, header in enumerate(["Country", "GDP Per Capita (USD)"], 1):
    cell = ws4.cell(row=2, column=col, value=header)
    style_header(cell)

top20_pc = pc_latest.head(20).reset_index(drop=True)
for i, row in top20_pc.iterrows():
    r = i + 3
    ws4.row_dimensions[r].height = 22
    bg = LIGHT_BLUE if i % 2 == 0 else WHITE
    for col, val in enumerate([row["Country"], row["GDP Per Capita 2022 (USD)"]], 1):
        cell = ws4.cell(row=r, column=col, value=val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = thin_border()
        style_cell(cell, align="center")

# ── SAVE ────────────────────────────────────────────────────────────────────
wb_excel.save(filename)
print(f"\n✅ Report saved → {filename}")
print("📊 Sheets: Executive Summary | Top 10 Economies | GDP Growth | GDP Per Capita")